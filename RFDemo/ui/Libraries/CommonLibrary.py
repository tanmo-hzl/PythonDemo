import datetime
import json
import os
import random
import re
import time
import platform
import requests
import uuid
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pykeyboard
import configparser
from robot.libraries.BuiltIn import BuiltIn
from requests_toolbelt.multipart.encoder import MultipartEncoder
from requests.packages.urllib3.exceptions import InsecureRequestWarning

requests.packages.urllib3.disable_warnings(InsecureRequestWarning)


class CommonLibrary(object):
	def __init__(self):
		pass

	@staticmethod
	def __get_root_path():
		parent_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
		return parent_path

	def get_config_ini(self, env_name, config_file="config.ini"):
		config_path = os.path.join(self.__get_root_path(), config_file)
		conf = configparser.ConfigParser()
		conf.read(config_path, encoding="utf-8")
		options = conf.options(env_name)
		config_info = {}
		for item in options:
			config_info[item] = conf.get(env_name, item)
		return config_info

	@staticmethod
	def my_choose_file(locator, path):
		sl = BuiltIn().get_library_instance("SeleniumLibrary")
		# sl.choose_file(locator, path)
		print(f"Sending {os.path.abspath(path)} to browser.")
		sl.find_element(locator).send_keys(path)

	@staticmethod
	def ctrl_or_command_key():
		if platform.system() == "Darwin":
			return "COMMAND"
		return "CONTROL"

	@staticmethod
	def return_or_enter_key():
		if platform.system() == "Darwin":
			return "RETURN"
		return "RETURN"

	@staticmethod
	def clear_input_value_by_keyboard():
		k = pykeyboard.PyKeyboard()
		if platform.system() == "Darwin":
			k.press_keys(['Command', 'A'])
			k.press_key('DELETE')
		else:
			k.press_keys(['CONTROL', 'A'])
			k.press_key('DELETE')
		time.sleep(1)

	@staticmethod
	def get_random_code(length=5, haveNumber=False):
		if haveNumber:
			string = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789"
		else:
			string = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ"
		randomId = random.sample(string, length)
		print("".join(randomId))
		return "".join(randomId)

	@staticmethod
	def get_uuid_split():
		return str(uuid.uuid1()).split("-")[0]

	@staticmethod
	def create_dir_not_existed(dir_path):
		if not os.path.exists(dir_path):
			os.mkdir(dir_path)

	def save_file(self, file_name, data, project_name=None, env='tst03'):
		if project_name is None:
			file_dir_path = os.path.join(self.__get_root_path(), 'CaseData')
		else:
			file_dir_path = os.path.join(self.__get_root_path(), 'CaseData', "{}-{}".format(project_name.upper(), env.upper()))
			self.create_dir_not_existed(file_dir_path)
		new_data = None
		try:
			if type(data) is list:
				new_data = {"data": data}
			else:
				if not type(data) is dict:
					new_data = json.loads(data)
				else:
					new_data = data
		except Exception as e:
			print(e)
			new_data = data
		finally:
			print("data type = ", type(new_data))
			if type(new_data) is dict:
				file_path = os.path.join(file_dir_path, file_name+".json")
				with open(file_path, "w+", encoding="utf-8") as f:
					json.dump(new_data, f, indent=4)
			else:
				file_path = os.path.join(file_dir_path, file_name + ".txt")
				with open(file_path, "w+", encoding="utf-8") as f:
					f.write(str(new_data))
		print(file_path)

	def read_file(self, file_name, project_name=None, env='tst03'):
		"""
		:param env:
		:param file_name:
		:param project_name:
		:return:
		"""
		file_type = 1
		if project_name is not None:
			file_path = os.path.join(self.__get_root_path(), "{}-{}".format(project_name.upper(), env.upper()), file_name + '.txt')
		else:
			file_path = os.path.join(self.__get_root_path(), file_name+'.txt')

		if not os.path.exists(file_path):
			file_type = 2
			if project_name is not None:
				file_path = os.path.join(self.__get_root_path(), "{}-{}".format(project_name.upper(), env.upper()), file_name + '.json')
			else:
				file_path = os.path.join(self.__get_root_path(), file_name+'.json')
		if not os.path.exists(file_path):
			return None
		print(file_path)
		with open(file_path, "r", encoding="utf-8") as f:
			if file_type == 1:
				data = f.read()
			else:
				data = json.load(f)
		return data

	@staticmethod
	def get_en_month(month, sort_month=False):
		en_month = "January,February,March,April,May,June,July,August,September,October,November,December"
		if sort_month:
			en_month = "Jan,Feb,Mar,Apr,May,Jun,Jul,Aug,Sep,Oct,Nov,Dec"
		en_months = en_month.split(",")
		return en_months[int(month)-1]

	@staticmethod
	def get_json_value(data, *args):
		"""
		根据传入的key值，返回 data,key|2 表示去list下标为2的key的值
		:param data:
		:param args:
		:return:
		"""
		try:
			for arg in args:
				if type(data) is list:
					if args.count("|"):
						num = re.findall(re.compile(r"\d"), arg)
						if not num:
							_num = 0
						else:
							_num = int(num[0])
						data = data[_num]
						ag = re.sub(r"\W|\d", "", arg)
						data = data[ag]
					else:
						data = data[0][arg]
				else:
					data = data[arg]
			return data
		except Exception as e:
			print(e)
			return None

	@staticmethod
	def get_all_files_in_folder(folder_name):
		all_files = []
		dir_or_files = os.listdir(folder_name)
		for dir_files in dir_or_files:
			if not os.path.isdir(os.path.join(folder_name, dir_files)):
				all_files.append(dir_files)
		return all_files

	def remove_download_file_by_part_name(self, file_name, is_default_downloads_path=True, custom_path=None):
		"""
		split filename by *, then check file startswith and endswith the split text
		:param file_name: aaa*.xlsx
		:param is_default_downloads_path:
		:param custom_path:
		:return:
		"""
		if is_default_downloads_path:
			downloads_path = self.get_default_downloads_path()
		else:
			downloads_path = custom_path
		if file_name.count("*"):
			names = file_name.split("*")
			all_files = self.get_all_files_in_folder(downloads_path)
			is_existed = False
			for file in all_files:
				if file.startswith(names[0]) and file.endswith(names[1]):
					is_existed = True
					file_path = os.path.join(downloads_path, file)
					os.remove(file_path)
					print("Remove file success! {}".format(file_path))
			if not is_existed:
				print("File not exited! {}".format(file_name))
		else:
			self.remove_download_file_if_existed(file_name)

	def remove_download_file_if_existed(self, file_name, is_default_downloads_path=True, custom_path=None):
		if is_default_downloads_path:
			downloads_path = self.get_default_downloads_path()
		else:
			downloads_path = custom_path
		file_path = os.path.join(downloads_path, file_name)
		if os.path.exists(file_path):
			os.remove(file_path)
			print("Remove file success! {}".format(file_path))
		else:
			print("File not exited! {}".format(file_path))

	def wait_until_file_download_by_part_name(self, file_name, time_out=5, is_default_downloads_path=True, custom_path=None):
		if is_default_downloads_path:
			downloads_path = self.get_default_downloads_path()
		else:
			downloads_path = custom_path
		if file_name.count("*"):
			names = file_name.split("*")
			start_time = time.time()
			file_path = None
			while True:
				all_files = self.get_all_files_in_folder(downloads_path)
				is_file_exist = False
				for file in all_files:
					if file.startswith(names[0]) and file.endswith(names[1]):
						is_file_exist = True
						file_path = os.path.join(downloads_path, file)
						break
				if is_file_exist:
					print("Get file success! {}".format(file_path))
					break
				end_time = time.time()
				if end_time - start_time > time_out:
					print("Get file time out! {}".format(file_name))
					break
				time.sleep(1)
			if file_path is not None:
				return True, file_path
			else:
				return False, None
		else:
			self.wait_until_file_download(file_name, time_out)

	def wait_until_file_download(self, file_name, time_out=5, is_default_downloads_path=True, custom_path=None):
		if is_default_downloads_path:
			file_path = os.path.join(self.get_default_downloads_path(), file_name)
		else:
			file_path = os.path.join(custom_path, file_name)
		start_time = time.time()
		while True:
			if os.path.exists(file_path):
				print("Get file success! {}".format(file_path))
				break
			else:
				time.sleep(1)
			end_time = time.time()
			if end_time - start_time > time_out:
				print("Get file time out! {}".format(file_path))
				break
		if os.path.exists(file_path):
			return True, file_path
		else:
			return False, None

	def read_download_excel(self, file_name, is_default_downloads_path=True, custom_path=None):
		if is_default_downloads_path:
			file_path = os.path.join(self.get_default_downloads_path(), file_name)
		else:
			file_path = os.path.join(custom_path, file_name)
		wb = load_workbook(file_path)
		all_sheet = wb.sheetnames
		sheet_name = all_sheet[0]
		sheet = wb[sheet_name]
		max_row = sheet.max_row
		max_column = sheet.max_column
		get_end = f'{get_column_letter(max_column)}{max_row}'
		start = "A1"
		end = get_end
		result = []
		for row_cell in sheet[start:end]:
			rows = []
			for cell in row_cell:
				if cell.value is None:
					rows.append("")
				else:
					rows.append(cell.value)
			result.append(rows)
		return result

	@staticmethod
	def get_default_downloads_path():
		user_path = os.path.expanduser('~')
		default_downloads_path = os.path.join(user_path, "Downloads")
		return default_downloads_path

	def create_custom_downloads_directory(self, dir_name="EA-0"):
		base_path = self.get_default_downloads_path()
		downloads_directory = os.path.join(base_path, dir_name)
		if not os.path.exists(downloads_directory):
			os.mkdir(downloads_directory)
			print("create {} success!".format(downloads_directory))
		else:
			print("{} is existed".format(downloads_directory))
		return downloads_directory

	@staticmethod
	def remove_custom_downloads_directory(downloads_directory):
		if os.path.exists(downloads_directory):
			os.rmdir(downloads_directory)
			print("remove {} success!".format(downloads_directory))
		else:
			print("{} not existed".format(downloads_directory))

	@staticmethod
	def check_string_contains(base_str, value):
		if value in base_str:
			return True
		else:
			return False

	def get_ea_fixed_element(self, file_name):
		file_path = os.path.join(self.__get_root_path(), "CaseData", "MP", "EA", file_name)
		with open(file_path, "r", encoding="utf-8") as f:
			data = json.load(f)
		return data

	def get_random_img_by_project_name(self, project_name="MP", quantity=1):
		img_dir_path = os.path.join(self.__get_root_path(), "CaseData", project_name, "IMG")
		img_list = os.listdir(img_dir_path)
		if quantity == 1:
			img_path = os.path.join(img_dir_path, random.choice(img_list))
			print(img_path)
			return img_path
		else:
			img_path = []
			images = random.sample(img_list, quantity)
			for item in images:
				img_path.append(os.path.join(img_dir_path, item))
			return img_path

	def check_page_date_in_filter_date_range(self, page_date, filter_date_range, purchased_within=None):
		result_date = self.format_page_date(page_date)
		start_date = self.format_filter_date(filter_date_range[0])
		end_date = self.format_filter_date(filter_date_range[1])
		message = "{{}},{}, page date {} {{}}in filter date range {} to {}".format(purchased_within, str(result_date)[:10], str(start_date)[:10], str(end_date)[:10])
		if start_date <= result_date <= end_date:
			print(message.format("Pass", ""))
		else:
			assert False, message.format("Fail", "not ")

	def check_filter_date_range_contains_page_date(self, filter_date_range, page_date_range):
		filter_start_date = self.format_filter_date(filter_date_range[0])
		filter_end_date = self.format_filter_date(filter_date_range[1])
		page_start_date = self.format_page_date(page_date_range[0])
		page_end_date = self.format_page_date(page_date_range[1])
		message = "{{}},filter_date_range {}-{} {{}}contains page_date_range {}-{}".\
			format(filter_start_date[:10], filter_end_date[:10], page_start_date[:10], page_end_date[:10])
		if filter_start_date <= page_start_date <= filter_end_date or filter_start_date <= page_end_date <=filter_end_date:
			assert True, message.format("Pass", "")
		else:
			assert False,message.format("Fail", "not ")

	@staticmethod
	def format_page_date(page_date):
		temp_result_date = page_date.split("/")
		result_date = "{}-{}-{}".format(temp_result_date[2], temp_result_date[0], temp_result_date[1])
		result_date = datetime.datetime.strptime(result_date, "%Y-%m-%d")
		return str(result_date)

	@staticmethod
	def format_filter_date(filter_date):
		filter_date = datetime.datetime.strptime(filter_date[:10], "%Y-%m-%d")
		return str(filter_date)

	def get_date_range_by_purchased_within(self, purchased_within):
		date_range = []
		within_list = ["All Time", "Today", "Yesterday", "Past 7 Days", "Past 30 Days", "Past 6 Months", "Past Year"]
		within_list_upper = []
		for item in within_list:
			within_list_upper.append(item.upper())
		index = -1
		for i in range(len(within_list_upper)):
			if within_list_upper[i].count(purchased_within.upper()):
				index = i

		if index == -1:
			assert False, "Fail, {} not existed".format(purchased_within)
		if index == 0:
			date_range.append(self.get_date_before_days(-100*365))
			date_range.append(self.get_date_before_days())
		if index == 1:
			date_range.append(self.get_date_before_days())
			date_range.append(self.get_date_before_days())
		if index == 2:
			date_range.append(self.get_date_before_days(-1))
			date_range.append(self.get_date_before_days(-1))
		if index == 3:
			date_range.append(self.get_date_before_days(-7))
			date_range.append(self.get_date_before_days())
		if index == 4:
			date_range.append(self.get_date_before_months(-1))
			date_range.append(self.get_date_before_days())
		if index == 5:
			date_range.append(self.get_date_before_days(-30*6))
			date_range.append(self.get_date_before_days())
		if index == 6:
			date_range.append(self.get_date_before_days(-365))
			date_range.append(self.get_date_before_days())
		return str(date_range)

	@staticmethod
	def get_date_before_days(days=0):
		need_date = (datetime.datetime.now() + datetime.timedelta(days=days)).strftime("%Y-%m-%d")
		return str(need_date)

	@staticmethod
	def get_date_before_months(months=-1):
		now = str(datetime.datetime.now().strftime("%Y-%m-%d")).split("-")
		year, month, day = now
		b_year, b_month = [months // 12, months % 12]
		if int(month) + b_month <= 0:
			month = 12 + int(month) + b_month
			year = int(year) - 1
		elif int(month) + b_month > 12:
			month = int(month) + b_month - 12
			year = int(year) + 1
		else:
			month = int(month) + b_month
		year = int(year) + b_year
		new_date = "{}-{}-{}".format(year, str(month).rjust(2, "0"), day)
		return str(datetime.datetime.strptime(new_date, "%Y-%m-%d"))

	@staticmethod
	def format_page_money(page_money):
		page_money = page_money.replace(",", "")
		money = re.findall(r"\d+.\d+", page_money)
		return float(money[0])

	@staticmethod
	def get_ea_listing_inventory(info):
		if info == "Out of stock":
			inventory = 0
		else:
			if info.count("\n"):
				inventory = (info.split("\n")[0]).split(" ")[0]
			else:
				inventory = info.split(" ")[0]
		return inventory

	@staticmethod
	def get_status_code_by_key(key):
		status_code = {"pending return": 11000, "refund rejected": 19100, "returned": 18000, "return canceled": 19000,
		               "pending refund": 16000, "dispute opened": 30000, "dispute in process": 30100,
		               "offer made": 31000, "offer rejected": 31100, "dispute resolved": 34000,
		               "dispute escalated": 32000, "escalation under review": 32100, "dispute canceled": 35000}
		return status_code.get(key.lower())

	def compare_sort_data(self, is_sort_down, first_data, last_data, data_type=None, data_name=""):
		"""

		:param is_sort_down: true/false
		:param first_data: 1/12/2022,$1,023.12,TPH
		:param last_data: 1/12/2022,$1,023.12,TPH
		:param data_type:  date,money,inventory,status,None=string
		:param data_name:
		:return:
		"""
		if data_type is None:
			data_type = "string"
		if data_type.upper() == "DATE":
			new_first_data = str(self.format_page_date(first_data))[:10]
			new_last_data = str(self.format_page_date(last_data))[:10]
		if data_type.upper() == "MONEY":
			new_first_data = self.format_page_money(first_data)
			new_last_data = self.format_page_money(last_data)
		if data_type.upper() == "INVENTORY":
			new_first_data = int(self.get_ea_listing_inventory(first_data))
			new_last_data = int(self.get_ea_listing_inventory(last_data))
			first_data = first_data.replace("\n", " ")
			last_data = last_data.replace("\n", " ")
		if data_type.upper() == "STATUS":
			new_first_data = self.get_status_code_by_key(first_data)
			new_last_data = self.get_status_code_by_key(last_data)
		if data_type.upper() == "STRING":
			new_first_data = first_data
			new_last_data = last_data

		if is_sort_down:
			if new_first_data < new_last_data:
				msg = 'Fail, {}, {}[{}, {}] {}sort down.'.format(data_name, data_type, first_data, last_data, "not " if is_sort_down else "")
				assert False, msg
			else:
				msg = 'Pass, {}, {}[{}, {}] sort down.'.format(data_name, data_type, first_data, last_data)
				print(msg)
		else:
			if new_first_data > new_last_data:
				msg = 'Fail, {}, {}[{}, {}] {}sort up.'.format(data_name, data_type, first_data, last_data, "not " if not is_sort_down else "")
				assert False, msg
			else:
				msg = 'Pass, {}, {}[{}, {}] sort up.'.format(data_name, data_type, first_data, last_data)
				print(msg)


if __name__ == '__main__':
	cc = CommonLibrary()
	a = cc.get_date_before_months()
	print(a)


