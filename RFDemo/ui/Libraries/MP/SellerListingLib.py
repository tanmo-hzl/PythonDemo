import datetime
import json
import random
import re
import time
import uuid
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

use_pro_type = True


class SellerListingLib(object):
	def __init__(self):
		self.root_path = self.__get_root_path()
		self.listing_status_list = ["Prohibited", "Pending Review", "Expired", "Out of stock", "Archived", "Active",
		                            "Draft", "Inactive", "Suspended"]
		self.categories = ['Frames', 'Supplies', 'Yarn', 'Needlework', 'Crochet', 'Knitting', 'Storage', 'Crafts',
		                   'Baking', 'Supplies', 'Beads', 'Jewelry', 'Crafts', 'Decor', 'Craft', 'Machines',
		                   'Fabric', 'Shop', 'Floral', 'Kids', 'Home', 'Office', 'Party', 'Teacher', 'Supplies',
		                   'Wedding']
		self.numberTypes = ["UPC", "UPC-A", "GS1-QRcode", "JAN", "EAN-8", "ISBN", "ASIN"]
		self.over_shipping_rate = {
			"0": "DefaultRate", "1": "OverrideRate"
		}
		self.custom_returns = {
			"0": "DefaultReturn", "1": "NoReturn", "2": "30Days @NeedItems", "3": "30Days @NoItems",
			"4": "60Days @NeedItems", "5": "60Days @NoItems",
		}

	@staticmethod
	def __get_root_path():
		abspath = os.path.dirname(os.path.abspath(__file__))
		root_path = os.path.dirname(os.path.dirname(abspath))
		return root_path

	@staticmethod
	def get_random_data():
		return str(uuid.uuid1()).split("-")[0]

	def get_random_img_path(self):
		img_dir_path = os.path.join(self.root_path, "CaseData", "MP", "IMG")
		img_list = os.listdir(img_dir_path)
		img_path = os.path.join(img_dir_path, random.choice(img_list))
		print(img_path)
		return img_path

	def get_listing_body(self, is_new=True, variants=False, variant_quantity=1,
	                     overShippingRate=None, customReturn=None):
		"""
		new title: "AU Single 220111x034750 DefaultRate DefaultReturn"
		:param is_new:
		:param variants:item have variants or not
		:param variant_quantity:
		:param overShippingRate:
		:param customReturn: 0:default, 1-no return;  2-30 days,don't need return items;
				3-30 days,need return items;  4-60 days,don't need return items;  5-60 days,need return items
		:return:
		"""
		categories = self.categories

		numberType = random.choice(self.numberTypes)
		colors = ["Orange", "Gold", "Yellow", "Green", "Purple", "Pink", "Silver", "Grey", "Black", "Brown", "White",
		          "Red", "Blue", "Clear"]
		variant_type = ["Size", "Color", "Count"]
		variant_colors = ["Red", "Orange", "Yellow", "Green", "Blue", "Pink", "Purple", "Grey", "Black", "Brown",
		                  "Gold", "Silver", "White", "Clear", "Multicolor"]
		other_color = "None of these work? Create a new one!"
		color = random.choice(colors)
		uu_code = "".join(str(uuid.uuid1()).split('-'))[:16]
		random_code = datetime.datetime.now().strftime("%y%m%dx%H%M%S")
		attributes = []
		if variant_quantity == 1:
			attributes_quantity = 2
		elif variant_quantity == 2:
			attributes_quantity = 5
		else:
			attributes_quantity = 11
		for i in range(attributes_quantity):

			attributes.append({"numberType": numberType,
			                   "number": self.get_global_trade_item_number_by_type(numberType),
			                   "sku": "".join(str(uuid.uuid1()).split('-'))[:16]})
			if overShippingRate is None:
				overShippingRate = str(random.randint(0, 1))

			if customReturn is None:
				customReturn = str(random.randint(0, 5))
		pro_type = " -{}@".format(random.randint(1, 6)) if use_pro_type else ""
		if is_new:
			if variants:
				title = "AU Variants {} @{} @{}{}" \
					.format(random_code, self.over_shipping_rate.get(overShippingRate), self.custom_returns.get(customReturn), pro_type)
			else:
				title = "AU Single {} @{} @{}{}" \
					.format(random_code, self.over_shipping_rate.get(overShippingRate), self.custom_returns.get(customReturn), pro_type)
		else:
			title = "AU {} @{} @{}{} ".format(random_code, self.over_shipping_rate.get(overShippingRate), self.custom_returns.get(customReturn), pro_type)
		body = {
			"title": title,
			"subTitle": ["Single", "Variants"],
			"category": random.choice(categories),
			"sellerSku": "{}".format(uu_code),
			"bandName": "Auto",
			"manufacturer": "Manufacturer {}".format(random_code),
			"recommendedAge": random.randint(1, 4),
			"description": "Item Description {}".format(datetime.datetime.now()),
			"tags": ["Auto", "Test" if is_new else "Change", "Variant" if variants else "Single"],
			"fromDate": (datetime.datetime.now() + datetime.timedelta(days=random.randint(0, 1))).strftime(
				"%Y-%m-%d"),
			"toDate": (datetime.datetime.now() + datetime.timedelta(days=30)).strftime("%Y-%m-%d"),
			"endDate": False if random.randint(1, 10) > 5 else True,
			"photos": [self.get_random_img_path()],
			"video": None,
			"variants": {
				"type": variant_type[0:variant_quantity],
				"Size": ["Small", "Normal", "Large"],
				"Count": ["Half dozen", "Six"],
				"Color": [random.choice(variant_colors), other_color],
				"otherColor": "Other",
				"quantity": "{}".format(random.randint(100, 999)),
				"price": "{}.{}".format(random.randint(0, 9), random.randint(10, 99)),
				"unVisible": random.randint(0, attributes_quantity),
				"attributes": attributes,
				"volumes": [1, 1, 2, 1, 2]
			},
			"singleSku": {
				"price": "{}.{}".format(random.randint(0, 9), random.randint(10, 99)),
				"quantity": "{}".format(random.randint(20, 999)),
				"attributes": {
					"color": color,
					"colorName": "{} {}".format(color, random_code),
					"numberType": numberType,
					"number": self.get_global_trade_item_number_by_type(numberType)},
				"volumes": [1, 1, 2, 1, 2]},
			"shippingPolicy": False if random.randint(1, 10) > 5 else True,
			"overShippingRate": {
				"status": False if overShippingRate == "0" else True,
				"standardFree": False if random.randint(1, 10) > 5 else True,
				"standard": "{}.{}".format(random.randint(1, 3), random.randint(10, 99)),
				"expedited": "{}.{}".format(random.randint(3, 6), random.randint(10, 99)),
				"freight": "{}.{}".format(random.randint(6, 9), random.randint(10, 99))
			},
			"customReturn": {
				"status": False if customReturn == "0" else True,
				"returnPolicy": customReturn
			}
		}
		if is_new:
			if variants:
				body['singleSku'] = None
			else:
				body['variants'] = None
		if variants:
			print("variant_detail = ", body)
		else:
			print("single_detail = ", body)
		return body

	@staticmethod
	def get_global_trade_item_number_by_type(number_type):
		if number_type == "GS1-QRcode":
			return random.randint(10000000000000, 99999999999999)
		if number_type == "ISBN":
			return random.randint(1000000000009, 9999999999990)
		if number_type == "UPC-A":
			return random.randint(100000000000, 999999999999)
		if number_type == "UPC":
			return random.randint(100000000000, 999999999999)
		if number_type == "ASIN":
			return random.randint(1000000000, 9999999999)
		if number_type == "JAN":
			return random.randint(100000000, 999999999)
		if number_type == "EAN-8":
			return random.randint(10000000, 99999999)

	@staticmethod
	def get_import_data(res):
		import_data = res[5:]
		for data in import_data[::-1]:
			if data[23] == "DRAFT":
				import_data.remove(data)
			elif data[1] == "":
				import_data.remove(data)
		return import_data

	def get_listing_status_by_number(self, disabled_status, quantity=1):
		status = list(set(self.listing_status_list) - set(disabled_status))
		if quantity <= 1:
			return random.sample(status, 1)
		if quantity > len(status):
			return status
		else:
			return random.sample(status, int(quantity))

	@staticmethod
	def get_listing_inventory(info):
		if info == "Out of stock":
			inventory = 0
		else:
			if info.count("\n"):
				inventory = (info.split("\n")[0]).split(" ")[0]
			else:
				inventory = info.split(" ")[0]
		return inventory

	@staticmethod
	def check_inventory_value_equal(data1, data2):
		if data1.count("\n"):
			data1 = " ".join(data1.split("\n"))

		if data2.count("\n"):
			data2 = " ".join(data1.split("\n"))

		return data1 == data2

	@staticmethod
	def check_listing_have_variants(info):
		if info.count("\n"):
			return True
		else:
			return False

	def get_random_category(self):
		return random.choice(self.categories)

	def get_listing_fixed_element(self, file_name):
		file_path = os.path.join(self.root_path, "CaseData", "MP", "EA", file_name)
		with open(file_path, "r", encoding="utf-8") as f:
			data = json.load(f)
		return data

	@staticmethod
	def update_listing_value_by_keys(data, value, *args):
		"""
		:param data: json data
		:param value:
		:param args: keys
		:return:
		"""
		if len(args) == 0:
			return data
		elif len(args) == 1:
			data[args[0]] = value
			return data
		elif len(args) == 2:
			data[args[0]][args[1]] = value
			return data

	def read_listing_info_from_excel(self, file_name):
		if os.path.exists(file_name):
			file_path = file_name
		else:
			file_path = os.path.join(self.__get_root_path(), "CaseData", "MP", "EA", "{}.xlsx".format(file_name))
		wb = load_workbook(file_path)
		sheet = wb["Template"]
		max_row = sheet.max_row
		max_column = sheet.max_column
		get_end = f"{get_column_letter(max_column)}{max_row}"
		start = "A1"
		result = []
		for row_cell in sheet[start:get_end]:
			rows = []
			for cell in row_cell:
				if cell.value is None:
					rows.append("")
				else:
					rows.append(cell.value)
			result.append(rows)
		title = result[4]
		required_keys = []
		for i in range(len(result[3])):
			if result[3][i].count("*"):
				required_keys.append(title[i])
		listing_data = []
		if len(result) > 5:
			lst_data = result[5:]
			for item in lst_data:
				title_keys = {}
				for i in range(len(item)):
					title_keys[title[i]] = item[i]
				listing_data.append(title_keys)
		print(listing_data)
		return listing_data, required_keys

	def save_download_listings_file(self, listing_data, download_path=None, file_name="all_listings"):
		if download_path is None:
			file_path = os.path.join(self.__get_root_path(), "CaseData", "MP", "EA", "{}.xlsx".format(file_name))
		else:
			default_downloads_path = os.path.join(os.path.expanduser('~'), "Downloads")
			file_path = os.path.join(default_downloads_path, download_path, "{}.xlsx".format(file_name))
		if os.path.exists(file_path):
			os.remove(file_path)
		with open(file_path, "wb") as f:
			f.write(listing_data.content)
		return file_path

	def write_listing_info_to_excel(self, file_name, listing_data, download_path=None, update_title=True):
		if os.path.exists(file_name):
			file_path = file_name
		else:
			file_path = os.path.join(self.__get_root_path(), "CaseData", "MP", "EA", "{}.xlsx".format(file_name))
		if download_path is not None:
			default_downloads_path = os.path.join(os.path.expanduser('~'), "Downloads")
			new_file_path = os.path.join(default_downloads_path, download_path, "{}.xlsx".format(file_name))
			if os.path.exists(new_file_path):
				os.remove(new_file_path)
			wb = Workbook()
			wb.create_sheet("Template")
			wb.save(new_file_path)
		if not os.path.exists(file_path):
			print("No find file: ", file_path)
			return None
		if download_path is not None:
			file_path = new_file_path

		self.update_listing_file_info(file_path, listing_data, update_title)
		print(file_path)
		return file_path

	@staticmethod
	def update_listing_file_info(file_path, listing_data, update_title=True):
		wb = load_workbook(file_path)
		sheet_name = wb["Template"]
		excel_data = []
		for item_data in listing_data:
			excel_column_data = []
			for item in item_data:
				excel_column_data.append(item_data.get(item))
			excel_data.append(excel_column_data)

		if update_title:
			index = 1
			if len(listing_data) > 0:
				for item in listing_data[0]:
					sheet_name.cell(row=5, column=index).value = item
					index += 1

		row_index = 6
		for row_data in excel_data:
			column_index = 1
			for item in row_data:
				sheet_name.cell(row=row_index, column=column_index).value = str(item)
				column_index += 1
			row_index += 1
		wb.save(file_path)

	@staticmethod
	def update_import_listing_category(listing_data, new_category):
		for item in listing_data:
			item["category_path"] = new_category
		return listing_data

	def update_import_listing_data(self, listing_data, required_keys, **kwargs):
		"""

		:param listing_data:
		:param required_keys:
		:param kwargs:
		:return:
		"""
		need_update_key = ['price', 'quantity', 'item_width', 'item_length', 'item_height', 'item_weight']
		for item_data in listing_data:
			required_value = self.get_required_value(item_data.get("sku_type"))
			for item in item_data:
				if item in required_keys and (item_data[item] == "" or item in need_update_key):
					item_data[item] = required_value[item]

		for item_data in listing_data:
			if "data_status" in kwargs.keys() and kwargs["data_status"] is not None:
				if item_data.get("status") == "EXPIRED":
					item_data["available_from"] = (datetime.datetime.now() + datetime.timedelta(days=random.randint(0, 1))).strftime(
						"%Y-%m-%d")
					item_data["available_to"] = (datetime.datetime.now() + datetime.timedelta(days=random.randint(20, 30))).strftime(
						"%Y-%m-%d")

				if item_data.get("status") == "OUT OF STOCK":
					item_data["quantity"] = random.randint(100, 999)

				if kwargs["data_status"].upper() == "SUSPENDED":
					item_data["description"] = "Test Fuck, {}".format(datetime.datetime.now())
					item_data["brand_name"] = "Fuck"

				if kwargs["data_status"].upper() != "SUSPENDED":
					item_data["description"] = "Import Excel Date , {}".format(datetime.datetime.now())
					item_data["brand_name"] = "Auto Import"

				if kwargs["data_status"].upper() == "EXPIRED":
					item_data["available_from"] = (datetime.datetime.now() + datetime.timedelta(days=random.randint(-30, -10))).strftime(
						"%Y-%m-%d")
					item_data["available_to"] = (datetime.datetime.now() + datetime.timedelta(days=random.randint(-9, -5))).strftime(
						"%Y-%m-%d")

				if kwargs["data_status"].upper() != "EXPIRED":
					item_data["available_from"] = (datetime.datetime.now() + datetime.timedelta(days=random.randint(0, 1))).strftime(
						"%Y-%m-%d")
					item_data["available_to"] = (datetime.datetime.now() + datetime.timedelta(days=random.randint(20, 30))).strftime(
						"%Y-%m-%d")

				if kwargs["data_status"].upper() == "OUT OF STOCK":
					item_data["quantity"] = 0

				if kwargs["data_status"].upper() != "OUT OF STOCK":
					item_data["quantity"] = random.randint(100, 999)

		for item_data in listing_data:
			item_data_keys = item_data.keys()
			for key in kwargs.keys():
				if item_data.get("sku_type") in ["Parent", "Standalone"]:
					if key == "data_status" and kwargs.get("data_status") in ["Active", "Inactive", "Draft"]:
						item_data["status"] = kwargs.get("data_status").upper()
					if key == "listing_status" and kwargs.get("listing_status") is not None:
						item_data["status"] = kwargs.get("listing_status").upper()
				if key in item_data_keys:
					item_data[key] = kwargs.get(key)

		for item_data in listing_data:
			if item_data.get("sku_type") in ["Parent", "Standalone"]:
				listing_name = self.reset_import_listing_name_by_listing_data(item_data, False)
				item_data["item_name"] = listing_name
			if item_data.get("sku_type") == "Child":
				item_data["item_name"] = listing_name

		if "log" in kwargs.keys():
			print("listing_data= ", listing_data)
		return listing_data

	@staticmethod
	def check_required_key_not_null(listing_data, required_keys):
		all_required = True
		parent_can_None = ["global_trade_item_number_type", "global_trade_item_number", "price", "quantity",
		                   "item_width", "item_length", "item_height", "item_weight", "item_weight_uom"]
		child_can_None = ["brand_name"]
		for item_data in listing_data:
			if item_data.get("sku_type") == "Parent":
				if item_data.get("status") not in ["ACTIVE", "INACTIVE", "DRAFT"]:
					all_required = False
					break
				now_required_keys = list(set(required_keys) - set(parent_can_None))
				for item in item_data:
					if item in now_required_keys and item_data[item] == "":
						all_required = False
						break
			elif item_data.get("sku_type") == "Child":
				now_required_keys = list(set(required_keys) - set(child_can_None))
				for item in item_data:
					if item in now_required_keys and item_data[item] == "":
						all_required = False
						break
			else:
				if item_data.get("status") not in ["ACTIVE", "INACTIVE", "DRAFT"]:
					all_required = False
					break
				for item in item_data:
					if item in required_keys and item_data[item] == "":
						all_required = False
						break
		return all_required

	@staticmethod
	def compare_output_results(listing_data, err_listing_data):
		err_info = ""
		for i in range(len(listing_data)):
			listing_data_item = listing_data[i]
			err_listing_data_item = err_listing_data[i]
			for key in listing_data_item:
				if str(listing_data_item[key]) != str(err_listing_data_item[key]):
					err_info += key + ":" + err_listing_data_item[key] + ";"
		return err_info

	def get_required_value(self, sku_type="Standalone"):
		"""
		:param sku_type: Standalone/Parent/Child
		:return:
		"""
		global_trade_item_number_type = random.choice(self.numberTypes)
		global_trade_item_number = self.get_global_trade_item_number_by_type(global_trade_item_number_type)
		required_value = {'seller_sku': "".join(str(uuid.uuid1()).split('-'))[:16],
		                  'category_path': '',
		                  'global_trade_item_number_type': global_trade_item_number_type,
		                  'global_trade_item_number': global_trade_item_number,
		                  'item_name': '',
		                  'brand_name': 'Auto Import',
		                  'description': 'Import Excel Date {}'.format(datetime.datetime.now()),
		                  'tag_1': 'Import',
		                  'status': '',
		                  'image_url_1': '',
		                  'price': ''  '{}.{}'.format(random.randint(0, 9), random.randint(10, 99)),
		                  'quantity': random.randint(100, 999),
		                  'item_width': '{}.{}'.format(random.randint(1, 9), random.randint(10, 99)),
		                  'item_length': '{}.{}'.format(random.randint(1, 9), random.randint(10, 99)),
		                  'item_height': '{}.{}'.format(random.randint(1, 9), random.randint(10, 99)),
		                  'item_weight': '{}.{}'.format(random.randint(1, 9), random.randint(10, 99)),
		                  'item_weight_uom': 'lb'}
		if sku_type == "Parent":
			required_value["global_trade_item_number_type"] = ""
			required_value["global_trade_item_number"] = ""
			required_value["price"] = ""
			required_value["quantity"] = ""
			required_value["item_width"] = ""
			required_value["item_length"] = ""
			required_value["item_height"] = ""
			required_value["item_weight"] = ""
			required_value["item_weight_uom"] = ""
		if sku_type == "Child":
			required_value["brand_name"] = ""

		return required_value

	def create_import_new_listing_data(self, data1, data2, request_keys, status, data_status, data_type=0, quantity=1):
		"""

		:param data1: listing data no variants
		:param data2: listing data have variants
		:param request_keys: listing data request keys
		:param status: Draft,Active,Inactive
		:param data_status: Out of stock,Suspended,Expired
		:param data_type: 0 : no variants, 1 : have variants, 2: all in
		:param quantity: listing quantity
		:return: new listing data
		"""
		new_listing_data = []
		data1 = self.update_import_listing_data(data1, request_keys, data_status=data_status, listing_status=status, log=False)
		data2 = self.update_import_listing_data(data2, request_keys, data_status=data_status, listing_status=status, log=False)
		global_trade_item_number_type = random.choice(self.numberTypes)
		category_path = self.get_new_category_by_old()
		if int(data_type) == 0 or int(data_type) == 2:
			for i in range(int(quantity)):
				temp_data = {}
				listing = data1[0]
				for key in listing:
					temp_data[key] = listing[key]
				temp_data["sku_number"] = ""
				temp_data["manufacture_name"] = ""
				temp_data["tag_1"] = "Import"
				temp_data["seller_sku"] = "".join(str(uuid.uuid1()).split('-'))[:16]
				temp_data["global_trade_item_number_type"] = global_trade_item_number_type
				temp_data["global_trade_item_number"] = self.get_global_trade_item_number_by_type(global_trade_item_number_type)
				temp_data["item_name"] = self.reset_import_listing_name_by_listing_data(listing, True)
				temp_data["category_path"] = category_path
				new_listing_data.append(temp_data)
		if int(data_type) == 1 or int(data_type) == 2:
			parent_count = 0
			variant_listing = []
			for item in data2:
				if item.get("sku_type") == "Parent":
					parent_count += 1
					if parent_count > 1:
						break
					else:
						variant_listing.append(item)
				else:
					variant_listing.append(item)

			for i in range(int(quantity)):
				item_name = self.reset_import_listing_name_by_listing_data(variant_listing[0], True)
				parent_sku = "".join(str(uuid.uuid1()).split('-'))[:16]
				for item in variant_listing:
					temp_data = {}
					for key in item:
						temp_data[key] = item[key]
					temp_data["sku_number"] = ""
					temp_data["item_name"] = item_name
					temp_data["manufacture_name"] = ""
					temp_data["tag_1"] = "Import"
					temp_data["category_path"] = category_path
					if temp_data.get("sku_type") == "Parent":
						temp_data["seller_sku"] = parent_sku
						temp_data["global_trade_item_number_type"] = ""
						temp_data["global_trade_item_number"] = ""
					else:
						temp_data["seller_sku"] = "".join(str(uuid.uuid1()).split('-'))[:16]
						temp_data["parent_seller_sku"] = parent_sku
						temp_data["global_trade_item_number_type"] = global_trade_item_number_type
						temp_data["global_trade_item_number"] = self.get_global_trade_item_number_by_type(global_trade_item_number_type)
					new_listing_data.append(temp_data)
		return new_listing_data, category_path

	def reset_import_listing_name_by_listing_data(self, listing_data, is_new=False):
		sku_type = listing_data.get("sku_type")
		listing_name = listing_data.get("item_name")
		now_time = datetime.datetime.now().strftime("%y%m%dx%H%M%S")
		uu_code = str(uuid.uuid1()).split("-")[0]
		if not listing_name.count("Days") and not listing_name.count("Return"):
			if listing_data.get("override_shipping_rate") == "N":
				overRate = self.over_shipping_rate.get("0")
			else:
				overRate = self.over_shipping_rate.get("1")
			if listing_data.get("override_return_policy") == "N":
				customReturn = self.custom_returns.get("0")
			else:
				if listing_data.get("return_policy_option") == "NO_RETURN":
					customReturn = self.custom_returns.get("1")
				elif listing_data.get("return_policy_option") == "IN_30_DAYS":
					if listing_data.get("refund_only") == "N":
						customReturn = self.custom_returns.get("2")
					else:
						customReturn = self.custom_returns.get("3")
				else:
					if listing_data.get("refund_only") == "N":
						customReturn = self.custom_returns.get("4")
					else:
						customReturn = self.custom_returns.get("5")
			if is_new:
				action = "Create"
			else:
				action = "Update"
			pro_type = " -{}@".format(random.randint(1, 6)) if use_pro_type else ""
			if sku_type == "Standalone":
				listing_name = "AU Single Import-{} {} #{} @{} @{}{}".format(action, now_time, uu_code, overRate, customReturn, pro_type)
			else:
				listing_name = "AU Variants Import-{} {} #{} @{} @{}{}".format(action, now_time, uu_code, overRate, customReturn, pro_type)
		listing_name = update_listing_name(listing_name, is_new)
		return listing_name

	def check_listing_detail_is_update_after_import(self, listing_data, listing_detail):
		sku_type = listing_data[0].get("sku_type")
		import_listing = listing_data[0]
		actual_listing = listing_detail.get("listing")
		error_msg = self.get_different_listing_key_and_values(import_listing, actual_listing, sku_type)
		if sku_type in ['Parent']:
			sub_listings = listing_detail.get("subListings")
			child_import_listing = listing_data[random.randint(1,len(sub_listings))]
			for item in sub_listings:
				if item.get("sellerSkuNumber") == import_listing.get("seller_sku"):
					child_actual_listing = item
					error_msg = self.get_different_listing_key_and_values(child_import_listing, child_actual_listing, child_import_listing.get("sku_type"))
		if error_msg == "":
			return True, error_msg
		else:
			return False, error_msg

	@staticmethod
	def get_different_listing_key_and_values(import_listing, actual_listing, sku_type, error_msg=""):
		if sku_type in ['Standalone', 'Parent', 'Child']:
			if actual_listing.get("itemName") != import_listing.get("item_name"):
				error_msg += "itemName not update; "
			if actual_listing.get("longDescription") != import_listing.get("description"):
				error_msg += "longDescription not update; "
			if actual_listing.get("tags")[0] != import_listing.get("tag_1"):
				error_msg += "tags not update; "
			if actual_listing.get("onlineOnly").get("availableFromDate") != import_listing.get("available_from"):
				error_msg += "availableFromDate not update; "
			# if actual_listing.get("onlineOnly").get("availableToDate") != import_listing.get("available_to"):
			# 	error_msg += "availableToDate not update; "

		if sku_type in ['Standalone', 'Child']:
			if actual_listing.get("global_trade_item_number") != import_listing.get("globalTradeItemNumber"):
				error_msg += "globalTradeItemNumber not update; "
			if float(actual_listing.get("price")) != float(import_listing.get("price")):
				error_msg += "price not update; "
			if int(actual_listing.get("inventory")) != int(import_listing.get("quantity")):
				error_msg += "inventory not update; "
			if float(actual_listing.get("itemProperties").get("width")) != float(import_listing.get("item_width")):
				error_msg += "width not update; "
			if float(actual_listing.get("itemProperties").get("length")) != float(import_listing.get("item_length")):
				error_msg += "length not update; "
			if float(actual_listing.get("itemProperties").get("height")) != float(import_listing.get("item_height")):
				error_msg += "height not update; "
			if float(actual_listing.get("itemProperties").get("weight")) != float(import_listing.get("item_weight")):
				error_msg += "weight not update; "

		return error_msg

	def check_taxonomy_file_is_need_update(self, file_name="Taxonomy.json"):
		file_path = os.path.join(self.__get_root_path(), 'CaseData', 'MP', 'EA', file_name)
		is_need_update = True
		if os.path.exists(file_path):
			with open(file_path, "r", encoding="utf-8") as f:
				taxonomy = json.load(f)
			timestamp = taxonomy.get("timestamp")
			now_time = int(time.time())
			if now_time - timestamp < 20 * 60 * 60:
				is_need_update = False
		return is_need_update

	def save_taxonomy_file(self, data, file_name="Taxonomy.json"):
		file_path = os.path.join(self.__get_root_path(), 'CaseData', 'MP', 'EA', file_name)
		data["timestamp"] = int(time.time())
		with open(file_path, "w+", encoding="utf-8") as f:
			json.dump(data, f, indent=4)
		return file_path

	def get_new_category_by_old(self, old_category=None, file_name="Taxonomy.json"):
		file_path = os.path.join(self.__get_root_path(), 'CaseData', 'MP', 'EA', file_name)
		with open(file_path, "r", encoding="utf-8") as f:
			taxonomy = json.load(f)
		all_taxonomy = taxonomy.get("data")
		if old_category is not None:
			categories = old_category.split("//")
			last_category = categories[len(categories)-1]
			if last_category.count(" "):
				last_category = last_category.split(" ")[0]
			new_categories = []
			for item in all_taxonomy:
				if item.get("taxonomyName").count(last_category) or item.get("taxonomyPath").count(last_category):
					new_categories.append(item)
			if len(new_categories) > 0:
				new_category = random.choice(new_categories).get("taxonomyPath")
			else:
				new_category = random.choice(all_taxonomy).get("taxonomyPath")
		else:
			new_category = random.choice(all_taxonomy).get("taxonomyPath")
		return new_category


def update_listing_name(listing_name, is_new):

	if listing_name.count(" Import-Update "):
		if is_new:
			listing_name = listing_name.replace(" Import-Update ", " Import-Create ")

	if listing_name.count(" Import-Create "):
		if not is_new:
			listing_name = listing_name.replace(" Import-Create ", " Import-Update ")

	if listing_name.count(" Import "):
		if is_new:
			listing_name = listing_name.replace(" Import ", " Import-Create ")
		else:
			listing_name = listing_name.replace(" Import ", " Import-Update ")

	if not listing_name.count("Import"):
		if is_new:
			listing_name = listing_name.replace("AU ", "AU Import-Create ")
		else:
			listing_name = listing_name.replace("AU ", "AU Import-Update ")

	now_time = datetime.datetime.now().strftime("%y%m%dx%H%M%S")
	uu_code = str(uuid.uuid1()).split("-")[0]
	try:
		date = re.findall(r"\d{6}x\d{6}", listing_name)
		if is_new and len(date) > 0:
			listing_name = listing_name.replace(date[0], now_time)

		if len(date) == 0:
			listing_name = listing_name.replace("AU ", "AU {} ".format(now_time))

		if use_pro_type:
			pro_type_flag = re.findall(r"-\d@", listing_name)
			if len(pro_type_flag) == 0:
				listing_name = listing_name + " -{}@".format(random.randint(1, 6))
			if is_new and len(pro_type_flag) > 0:
				pro_type = "-{}@".format(random.randint(1, 6))
				listing_name = listing_name.replace(pro_type_flag[0], pro_type)

		uu_date = re.findall(r"#.{8}", listing_name)
		if len(uu_date) > 0:
			listing_name = listing_name.replace("{} ".format(uu_date[0]), "")
			listing_name = listing_name.replace("AU ", "AU #{} ".format(uu_code))
	except Exception as e:
		print(e)

	return listing_name


if __name__ == '__main__':
	lib = SellerListingLib()
	use_pro_type = True
	name = "AU Variants Import-Create #0608446a #0608446a @OverrideRate @30Days @NeedItems"
	n = update_listing_name(name, False)
	print(n)


