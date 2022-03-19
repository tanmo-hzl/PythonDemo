import os
import random

import openpyxl

file_path = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
file_path = os.path.join(file_path, 'CaseData')
file_path = os.path.join(file_path, 'API_Portal')
file_path = os.path.join(file_path, 'template.xlsx')
# file_path = os.path.join(file_path, 'import-create-active.xlsx')
numberSet = '01234567891'
seller_sku = ''
for i in range(14):
    seller_sku += random.choice(numberSet)

gtin = ''
for i in range(8):
    gtin += random.choice(numberSet)


catatory = 'root//Shop Categories//Frames//Shop By Frame Size'

media_url = 'https://imgproxy.qa.platform.michaels.com/XXTpt9MjeIKtJwhEo-W6WCFk1c9-K_Ll4qb5g7o5vco/aHR0cHM6Ly9zdGF0aWMucGxhdGZvcm0ubWljaGFlbHMuY29tL3RzdDAzLzYxNjUwNjU2MzUxNTM5MjAwMDAuanBlZw.jpeg'

excel = openpyxl.load_workbook(file_path)
print(excel.sheetnames)
sheet = excel['Template']
for i in range(6,1006):
    sheet.cell(row=i, column=1, value='Standalone')
    sheet.cell(row=i, column=2, value=str(int(seller_sku)+int(i)))
    sheet.cell(row=i, column=5, value=catatory)
    sheet.cell(row=i, column=6, value='EAN-8')
    sheet.cell(row=i, column=7, value=str(int(gtin)+int(i)))
    sheet.cell(row=i, column=8, value=f'0126 1000 listings test {i}')
    sheet.cell(row=i, column=9, value='apple')
    sheet.cell(row=i, column=13, value='aaa')
    sheet.cell(row=i, column=14, value='bbb')
    sheet.cell(row=i, column=24, value='ACTIVE')
    sheet.cell(row=i, column=27, value=media_url)
    sheet.cell(row=i, column=34, value='100')
    sheet.cell(row=i, column=35, value='100')
    sheet.cell(row=i, column=38, value='1')
    sheet.cell(row=i, column=39, value='1')
    sheet.cell(row=i, column=40, value='1')
    sheet.cell(row=i, column=41, value='1')
    sheet.cell(row=i, column=42, value='lb')

excel.save(file_path)
excel.close()


# excel = openpyxl.load_workbook(file_path)
# sheet = excel['Template']
#
# print(sheet['B5'].value)
#
# excel.close()


