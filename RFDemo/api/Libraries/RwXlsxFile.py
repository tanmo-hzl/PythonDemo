import copy
import os
from typing import Optional

import requests
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class RwXlsxFile(object):
    def __init__(self):
        pass

    @staticmethod
    def __get_root_path():
        parent_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        return parent_path

    def get_file_path_by_name(self, file_name, file_dir: Optional = None):
        if file_dir is not None:
            # dir_path = os.path.join(self.__get_root_path(), "CaseData", file_dir)
            dir_path = file_dir
            if not os.path.exists(dir_path):
                os.mkdir(dir_path)
        else:
            dir_path = os.path.join(self.__get_root_path(), "CaseData")
        file_path = os.path.join(dir_path, file_name)
        return file_path

    def load_file(self, file_name, file_dir: Optional = None):
        file_path = self.get_file_path_by_name(file_name, file_dir)
        f_path, f_name = os.path.split(file_path)
        name, ext = os.path.splitext(f_name)
        if ext in [".xlsx"]:
            if not os.path.exists(file_path):
                wb = Workbook()
                wb.save(file_path)
            f = load_workbook(file_path)
            return f
        else:
            return None

    def get_excel_sheets(self, file_name, file_dir: Optional = None):
        wb = self.load_file(file_name, file_dir)
        if wb is not None:
            sn = wb.sheetnames
            print(sn)
            return sn
        else:
            print("Not .xlsx file")
            return None

    @staticmethod
    def __get_active_sheet_name(wb, sheet_name: Optional = None):
        all_sheet = wb.sheetnames
        if sheet_name is None:
            sheet_name = all_sheet[0]
        else:
            if sheet_name not in all_sheet:
                wb.create_sheet(sheet_name)
        return sheet_name

    def write_excel(
        self, data, file_name, file_dir: Optional = None, sheet_name: Optional = None
    ):
        file_path = self.get_file_path_by_name(file_name, file_dir)
        wb = Workbook()
        try:
            sheet_name = self.__get_active_sheet_name(wb, sheet_name)
            sheet = wb[sheet_name]
            sheet.title = sheet_name
            for row_line in data:
                sheet.append(row_line)
            wb.save(file_path)
        except Exception as e:
            print(e)

    def read_excel(
        self,
        file_name,
        file_dir: Optional = None,
        sheet_name: Optional = None,
        start: Optional = None,
        end: Optional = None,
    ):
        wb = self.load_file(file_name, file_dir)
        sheet_name = self.__get_active_sheet_name(wb, sheet_name)
        sheet = wb[sheet_name]
        max_row = sheet.max_row
        max_column = sheet.max_column
        get_end = f"{get_column_letter(max_column)}{max_row}"
        start = start if start else "A1"
        end = end if end else get_end
        result = []
        for row_cell in sheet[start:end]:
            rows = []
            for cell in row_cell:
                if cell.value is None:
                    rows.append("")
                else:
                    rows.append(cell.value)
            result.append(rows)
        return result

    def get_par_category_info(self):
        d = self.read_excel("VariantTypeResult.xlsx", "MP", "mapping")
        categories = d[0][2:]
        print(categories)
        return categories

    def get_cate_mapping_info(self):
        categories = self.get_par_category_info()
        mapping = self.read_excel("VariantTypeResult.xlsx", "MP", "mapping1")
        mapping_v = mapping[0]
        mapping1 = mapping[2:]
        cate_mapping = {}
        for category_index in range(len(categories)):
            cate_mapping[categories[category_index]] = []
            for mapping1_item in mapping1:
                if mapping1_item[0] == categories[category_index]:
                    for index in range(1, len(mapping1_item)):
                        if mapping1_item[index] != "":
                            cate_mapping[categories[category_index]].append(
                                mapping_v[index] + "|" + "1"
                            )
                        else:
                            cate_mapping[categories[category_index]].append(
                                mapping_v[index] + "|" + "0"
                            )
        print(cate_mapping)
        return cate_mapping

    def get_mapping_values(self):
        values = self.read_excel("VariantTypeResult.xlsx", "MP", "values1")
        mapping_value = {}
        for value in values:
            mapping_value[value[0]] = []
            for index in range(1, len(value)):
                if value[index] != "":
                    mapping_value[value[0]].append(value[index])
        print(mapping_value)
        return mapping_value

    # def comparison_variant_type_results(self, env='tst03'):
    #     check_result = True
    #     base_file = self.get_file_path_by_name("cate_mapping.json", "MP-{}".format(env.upper()))
    #     results_file = self.get_file_path_by_name("page_cate_mapping.json", "MP-{}".format(env.upper()))
    #     with open(base_file, "r") as f:
    #         base_data = json.load(f)
    #     with open(results_file, "r") as f:
    #         results_data = json.load(f).get("data")[0]
    #     # print(base_data)
    #     # print(results_data)
    #     new_export = []
    #     title_export = ["title"]
    #     for item in base_data:
    #         item_data = base_data[item]
    #         item_export = [item]
    #         for new_item in item_data:
    #             new_items = new_item.split("|")
    #             if new_items[0] in results_data.get(item) and new_items[1] == "1":
    #                 item_export.append("TRUE")
    #             elif new_items[0] in results_data.get(item) and new_items[1] == "0":
    #                 check_result = False
    #                 item_export.append("FALSE")
    #             else:
    #                 item_export.append("")
    #             if new_items[0] not in title_export:
    #                 title_export.append(new_items[0])
    #         new_export.append(item_export)
    #     new_export.insert(0, title_export)
    #     self.write_excel(new_export, "VariantTypeCheckResults.xlsx", "MP-{}".format(env.upper()))
    #     return check_result

    # def comparison_variant_type_value_results(self, env='tst03'):
    #     check_result = True
    #     new_export = []
    #     base_file = self.get_file_path_by_name("mapping_value.json", "MP-{}".format(env.upper()))
    #     results_file = self.get_file_path_by_name("page_mapping_value.json", "MP-{}".format(env.upper()))
    #     with open(base_file, "r") as f:
    #         base_data = json.load(f)
    #     with open(results_file, "r") as f:
    #         results_data = json.load(f).get("data")[0]
    #     for item in base_data:
    #         item_base_data = base_data[item]
    #         item_export = [item]
    #         for item_data in item_base_data:
    #             if item_data in results_data.get(item):
    #                 item_export.append(item_data)
    #             else:
    #                 item_export.append("FALSE-{}".format(item_data))
    #                 check_result = False
    #         new_export.append(item_export)
    #     print(new_export)
    #     self.write_excel(new_export, "VariantTypeCheckValueResults.xlsx", "MP-{}".format(env.upper()))
    #     return check_result

    def split_sheet_with_color(
        self, file_name, file_dir: Optional = None, sheet_name: Optional = None
    ):
        wb = self.load_file(file_name, file_dir)
        sheet_name = self.__get_active_sheet_name(wb, sheet_name)
        sheet = wb[sheet_name]
        rows, cols = sheet.max_row, sheet.max_column
        common_lines = []
        color_lines = []
        for i in range(rows):
            col_line = []
            com_line = []
            for j in range(cols):
                ce = sheet.cell(row=i + 1, column=j + 1)
                if ce.fill.fgColor.rgb == "00000000" and ce.value is not None:
                    # "FFFF0000 is red"
                    com_line.append(ce.value)
                elif ce.value is not None:
                    col_line.append(ce.value)
            if com_line:
                common_lines.append(com_line)
            if col_line:
                color_lines.append(col_line)
        return common_lines, color_lines

    @staticmethod
    def get_all_taxonomies():
        request = requests.session()
        url = "https://mda.tst03.platform.michaels.com/api/developer/api/v1/listing/taxonomy"
        headers = {
            "Api-Key": "b103ac83120da792cb59a12e3715b4a9e5bb55913a13225d1c13875eb66134c40a1df1f4bb25fff3"
        }
        res = request.get(url, headers=headers)
        # print(res.json()["data"])
        return res.json()["data"]

    @staticmethod
    def get_category():
        request = requests.session()
        url = "https://mik.tst03.platform.michaels.com/api/sch/search/michaels/category"
        res = request.get(url)
        # print(res.json()["searchResults"]["items"])
        return res.json()["searchResults"]["items"]

    def copy_excel(
        self, file_name, file_dir: Optional = None, sheet_name: Optional = None
    ):
        wb = self.load_file(file_name, file_dir)
        sheet_name = self.__get_active_sheet_name(wb, sheet_name)
        sheet = wb[sheet_name]
        wb.copy_worksheet(sheet)
        file_path = self.get_file_path_by_name(file_name, file_dir)
        wb.save(file_path)

    def compare_category_path(self):
        excel_color_line = self.split_sheet_with_color(
            "Categories for Item Setup - Updated Version 12.09.21.xlsx", "MP", "Sheet0"
        )[1]
        all_category = self.get_category()
        print(len(all_category))
        category_paths = []
        for category in all_category:
            category_path = category["categoryPath"][6:].split("//")
            category_paths.append(category_path)
        inter = [i for i in category_paths if i in excel_color_line]
        excel_data = self.read_excel(
            "Categories for Item Setup - Updated Version 12.09.21.xlsx", "MP", "Sheet0"
        )
        excel_data_copy = copy.deepcopy(excel_data)
        print(len(excel_data))
        for data in excel_data:
            if data in category_paths:
                data.append("exist")
            else:
                data.append("not exist")
            if data[:4] in inter:
                data.append("need remove")
        for ui_path in category_paths:
            if ui_path not in excel_data_copy:
                while len(ui_path) <= 5:
                    ui_path.append("")
                ui_path.append("not exist")
                excel_data.append(ui_path)
        self.write_excel(excel_data, "Categories_path 01.17.22.xlsx", "MP")



if __name__ == "__main__":
    p = RwXlsxFile()
    # p.get_cate_mapping_info()
    # p.get_mapping_values()
    # print(p.comparison_variant_type_results())
    # print(p.get_par_category_info())
    # p.copy_excel("Categories for Item Setup - Updated Version 12.09.21.xlsx", "MP")
    # p.compare_category_path()

